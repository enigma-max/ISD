"""
============================================================
FoodPanda BD — Nationwide Restaurant Scraper → Excel
Covers all major cities + district-level grid across Bangladesh
============================================================
SETUP:
    !pip install requests openpyxl pandas

RUN:
    %run scraper.py
    # Output: /kaggle/working/foodpanda_bd_nationwide.xlsx
============================================================
"""

import requests
import time
import uuid
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG
# ============================================================
OUTPUT_PATH = "/kaggle/working/foodpanda_bd_nationwide.xlsx"
LIMIT_PER_LOCATION = 100   # max per API call batch
OFFSETS_PER_LOCATION = range(0, 200, 20)  # 0..180 → up to 200 per coord

# ============================================================
# ALL MAJOR BD CITIES / DISTRICT CENTRES
# Each tuple: (city_name, latitude, longitude)
# ============================================================
BD_LOCATIONS = [
    # ── Dhaka Division ──────────────────────────────────────
    ("Dhaka - Gulshan",        23.7925,  90.4078),
    ("Dhaka - Dhanmondi",      23.7461,  90.3742),
    ("Dhaka - Mirpur",         23.8223,  90.3654),
    ("Dhaka - Uttara",         23.8759,  90.3795),
    ("Dhaka - Motijheel",      23.7338,  90.4172),
    ("Dhaka - Mohammadpur",    23.7644,  90.3584),
    ("Dhaka - Wari",           23.7157,  90.4125),
    ("Narayanganj",            23.6238,  90.4998),
    ("Gazipur",                23.9999,  90.4203),
    ("Manikganj",              23.8634,  89.9880),
    ("Munshiganj",             23.5422,  90.5303),
    ("Narsingdi",              23.9221,  90.7153),
    ("Faridpur",               23.6070,  89.8429),
    ("Madaripur",              23.1641,  90.1921),
    ("Gopalganj",              23.0050,  89.8268),
    ("Tangail",                24.2513,  89.9167),
    ("Kishoreganj",            24.4449,  90.7766),
    ("Rajbari",                23.7578,  89.6436),
    ("Shariatpur",             23.2424,  90.4346),

    # ── Chittagong Division ──────────────────────────────────
    ("Chittagong - Agrabad",   22.3303,  91.8316),
    ("Chittagong - GEC",       22.3569,  91.8326),
    ("Chittagong - Halishahar",22.3750,  91.7861),
    ("Cox's Bazar",            21.4272,  92.0058),
    ("Comilla",                23.4607,  91.1809),
    ("Feni",                   23.0153,  91.3968),
    ("Lakshmipur",             22.9449,  90.8416),
    ("Noakhali",               22.8696,  91.0996),
    ("Chandpur",               23.2333,  90.6516),
    ("Brahmanbaria",           23.9570,  91.1115),
    ("Rangamati",              22.6522,  92.1965),
    ("Bandarban",              22.1953,  92.2184),
    ("Khagrachhari",           23.1193,  91.9847),

    # ── Sylhet Division ──────────────────────────────────────
    ("Sylhet",                 24.8949,  91.8687),
    ("Habiganj",               24.3745,  91.4156),
    ("Moulvibazar",            24.4826,  91.7774),
    ("Sunamganj",              25.0658,  91.3950),

    # ── Rajshahi Division ────────────────────────────────────
    ("Rajshahi",               24.3745,  88.6042),
    ("Bogura",                 24.8510,  89.3697),
    ("Pabna",                  24.0064,  89.2372),
    ("Sirajganj",              24.4507,  89.7001),
    ("Natore",                 24.4204,  88.9843),
    ("Naogaon",                24.7936,  88.9312),
    ("Chapainawabganj",        24.5965,  88.2786),
    ("Joypurhat",              25.1031,  89.0228),

    # ── Khulna Division ──────────────────────────────────────
    ("Khulna",                 22.8456,  89.5403),
    ("Jessore",                23.1667,  89.2167),
    ("Satkhira",               22.7185,  89.0705),
    ("Bagerhat",               22.6602,  89.7854),
    ("Narail",                 23.1724,  89.5120),
    ("Magura",                 23.4876,  89.4196),
    ("Chuadanga",              23.6402,  88.8418),
    ("Meherpur",               23.7621,  88.6318),
    ("Jhenaidah",              23.5447,  89.1525),
    ("Kushtia",                23.9015,  89.1212),

    # ── Barisal Division ─────────────────────────────────────
    ("Barisal",                22.7010,  90.3535),
    ("Bhola",                  22.6858,  90.6482),
    ("Patuakhali",             22.3596,  90.3298),
    ("Pirojpur",               22.5831,  89.9754),
    ("Jhalokati",              22.6404,  90.1987),
    ("Barguna",                22.1500,  90.1167),

    # ── Rangpur Division ─────────────────────────────────────
    ("Rangpur",                25.7439,  89.2752),
    ("Dinajpur",               25.6279,  88.6338),
    ("Kurigram",               25.8074,  89.6363),
    ("Gaibandha",              25.3288,  89.5288),
    ("Nilphamari",             25.9310,  88.8560),
    ("Lalmonirhat",            25.9923,  89.2847),
    ("Thakurgaon",             26.0318,  88.4616),
    ("Panchagarh",             26.3408,  88.5554),

    # ── Mymensingh Division ──────────────────────────────────
    ("Mymensingh",             24.7471,  90.4203),
    ("Jamalpur",               24.9375,  89.9378),
    ("Sherpur",                25.0204,  90.0152),
    ("Netrokona",              24.8710,  90.7268),
]

# ============================================================
# FOODPANDA API
# ============================================================
VENDOR_LIST_URL = "https://bd.fd-api.com/vendors-gateway/api/v1/pandora/vendors"

API_HEADERS = {
    "x-disco-client-id": "pd-microfrontend/web-acquisition",
    "X-FP-API-KEY": "volo",
    "User-Agent": "Mozilla/5.0",
}

# ============================================================
# HELPERS
# ============================================================

def safe_uuid() -> str:
    return str(uuid.uuid4())

def map_pricing(budget) -> str:
    try:
        return "৳" * min(int(budget), 3)
    except (TypeError, ValueError):
        return "N/A"

# ============================================================
# FETCH
# ============================================================

def fetch_for_location(city: str, lat: float, lng: float) -> list:
    items = []
    seen_codes = set()
    for offset in OFFSETS_PER_LOCATION:
        params = {
            "latitude":    lat,
            "longitude":   lng,
            "language_id": 1,
            "country":     "bd",
            "limit":       20,
            "offset":      offset,
        }
        try:
            resp = requests.get(VENDOR_LIST_URL, headers=API_HEADERS, params=params, timeout=15)
            resp.raise_for_status()
            batch = resp.json().get("data", {}).get("items", [])
            if not batch:
                break
            new = [r for r in batch if r.get("code") not in seen_codes]
            seen_codes.update(r.get("code") for r in new)
            items.extend(new)
        except Exception as e:
            print(f"    WARN: {e}")
        time.sleep(0.8)
    return items


def fetch_all_bangladesh() -> list:
    all_restaurants = {}   # keyed by vendor_code for dedup
    total_locations = len(BD_LOCATIONS)

    for idx, (city, lat, lng) in enumerate(BD_LOCATIONS, 1):
        print(f"[{idx:02d}/{total_locations}] {city} ({lat}, {lng}) ...", end=" ", flush=True)
        items = fetch_for_location(city, lat, lng)
        before = len(all_restaurants)
        for r in items:
            code = r.get("code")
            if code and code not in all_restaurants:
                r["_source_city"] = city   # tag which city found it
                all_restaurants[code] = r
        new_count = len(all_restaurants) - before
        print(f"{new_count} new  (total unique: {len(all_restaurants)})")

    return list(all_restaurants.values())

# ============================================================
# PARSE INTO 4 TABLES
# ============================================================

def parse_all(restaurants: list) -> dict:
    rest_rows     = []
    location_rows = []
    rating_rows   = []
    discount_rows = []

    for r in restaurants:
        restaurant_id = safe_uuid()
        cuisines      = r.get("cuisines") or []
        cuisine_str   = ", ".join(c["name"] for c in cuisines if c.get("name")) or "Unknown"

        # ── restaurant ──────────────────────────────────────
        rest_rows.append({
            "restaurant_id": restaurant_id,
            "name":          r.get("name", ""),
            "description":   r.get("description") or "",
            "pricing":       map_pricing(r.get("budget", 0)),
            "cuisine_type":  cuisine_str,
            "source_city":   r.get("_source_city", ""),
            "cover_url":     r.get("hero_image") or "",
            "logo_url":      r.get("logo") or "",
        })

        # ── location ────────────────────────────────────────
        lat = r.get("latitude")
        lng = r.get("longitude")
        if lat and lng:
            location_rows.append({
                "location_id":     safe_uuid(),
                "restaurant_id":   restaurant_id,
                "restaurant_name": r.get("name", ""),
                "latitude":        lat,
                "longitude":       lng,
            })

        # ── rating ──────────────────────────────────────────
        raw_rating = r.get("rating")
        if raw_rating:
            rating_rows.append({
                "rating_id":       safe_uuid(),
                "restaurant_id":   restaurant_id,
                "restaurant_name": r.get("name", ""),
                "rating":          round(float(raw_rating)),
                "review_count":    r.get("review_number", 0),
            })

        # ── discount ────────────────────────────────────────
        for d in (r.get("discounts_info") or []):
            value = d.get("value")
            if not value:
                continue
            discount_rows.append({
                "discount_id":     safe_uuid(),
                "restaurant_id":   restaurant_id,
                "restaurant_name": r.get("name", ""),
                "name":            d.get("message") or d.get("name") or "Discount",
                "discount":        float(value),
                "start_date":      d.get("start_date") or "",
                "end_date":        d.get("end_date")   or "",
            })

    return {
        "Restaurant": pd.DataFrame(rest_rows),
        "Location":   pd.DataFrame(location_rows),
        "Rating":     pd.DataFrame(rating_rows),
        "Discount":   pd.DataFrame(discount_rows),
    }

# ============================================================
# EXCEL STYLING
# ============================================================

HEADER_FILL = PatternFill("solid", fgColor="E63946")
ALT_FILL    = PatternFill("solid", fgColor="FFF1F2")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT   = Font(size=10)
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left",   vertical="center")
THIN        = Side(style="thin", color="DDDDDD")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_sheet(ws, df):
    for col_idx, col_name in enumerate(df.columns, 1):
        cell           = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        cell.border    = BORDER

    for row_idx in range(2, ws.max_row + 1):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx in range(1, ws.max_column + 1):
            cell           = ws.cell(row=row_idx, column=col_idx)
            cell.fill      = fill
            cell.font      = BODY_FONT
            cell.alignment = LEFT
            cell.border    = BORDER

    for col_idx, col_name in enumerate(df.columns, 1):
        vals    = df.iloc[:, col_idx - 1].astype(str)
        max_len = max(len(str(col_name)), vals.map(len).max() if not vals.empty else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    ws.freeze_panes             = "A2"
    ws.row_dimensions[1].height = 22

# ============================================================
# MAIN
# ============================================================

def run():
    print("=" * 60)
    print("  FoodPanda BD — Nationwide Scraper")
    print(f"  Covering {len(BD_LOCATIONS)} locations across all 8 divisions")
    print("=" * 60)

    print("\n[1/3] Fetching restaurants across Bangladesh ...")
    restaurants = fetch_all_bangladesh()
    print(f"\n  Grand total unique restaurants: {len(restaurants)}\n")

    print("[2/3] Parsing into 4 tables ...")
    tables = parse_all(restaurants)
    for name, df in tables.items():
        print(f"  {name}: {len(df)} rows")

    print(f"\n[3/3] Writing to {OUTPUT_PATH} ...")
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        for sheet_name, df in tables.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            style_sheet(writer.sheets[sheet_name], df)

    print(f"\nDone!  Saved → {OUTPUT_PATH}")

if __name__ == "__main__":
    run()